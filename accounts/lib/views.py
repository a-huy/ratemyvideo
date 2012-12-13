# File to store helper functions for views
import tempfile
import os
import collections
import datetime

from openpyxl.workbook import Workbook
from openpyxl.writer.excel import ExcelWriter
from openpyxl.cell import get_column_letter

# Export a list of users into a .xlsx spreadsheet
def export_users_list_to_xlsx(users):
    workbook = Workbook()
    out_file = r'users-list-' + str(now())
    ws = workbook.worksheets[0]
    ws.title = 'Users List'

    # Create column labels
    headers = ['First Name', 'Middle Name', 'Last Name', 'Email', 'Age', 'Gender',
               'Location', 'Rated', 'Balance', 'Earned', 'Joined']
    for argi in xrange(1, len(headers) + 1):
        col = get_column_letter(argi)
        cell =  ws.cell('%s1' % col)
        cell.value = headers[argi - 1]
        cell.style.font.bold = True
        if argi < 4: ws.column_dimensions[col].width = 12
        elif argi == 4: ws.column_dimensions[col].width = 25
        elif argi == 7: ws.column_dimensions[col].width = 20
        else: ws.column_dimensions[col].width = 10

    for argi in xrange(2, len(users) + 2):
        name_parts = users[argi - 2].real_name.split(' ')
        users[argi - 2].first_name = name_parts[0]
        users[argi - 2].last_name = name_parts[-1]
        users[argi - 2].middle_name = name_parts[1] if len(name_parts) > 2 else ''
        users[argi - 2].joined = users[argi - 2].created_date
        users[argi - 2].gender = users[argi - 2].gender[0].upper()
        for argt in xrange(1, len(headers) + 1):
            col = get_column_letter(argt)
            attr = re.sub(r' ', '_', headers[argt - 1].lower())
            cell = ws.cell('%s%s' % (col, argi))
            cell.value = getattr(users[argi - 2], attr)
            cell.style.font.size = 8

    fd, fn = tempfile.mkstemp()
    os.close(fd)
    workbook.save(filename = fn)
    fh = open(fn, 'rb')
    resp = fh.read()
    fh.close()
    response = HttpResponse(resp, mimetype='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment; filename="%s.xlsx"' % out_file
    return response

# Takes a list of datetime objects and returns a list of [date, count]
def count_by_date(items):
    dates_list = []
    counts = collections.Counter([datetime.datetime(month=x.month, \
        day=x.day, year=x.year) for x in items])
    map(lambda x: dates_list.append(['%s-%s-%s' % (x.month, x.day, x.year),
        counts[x]]), sorted(list(counts)))
    return dates_list

# takes a list of locations as strings and returns a list of [state, count]
def count_by_state(items):
    states_list = []
    counts = collections.Counter([str(x.split(', ')[-1]) for x in items])
    map(lambda x: states_list.append([x, counts[x]]), list(counts))
    return states_list

# takes a list of tuples (date, reward) and returns a list of [date, sum]
def sum_by_date(items):
    dates_dict = { }
    for pair in items:
        date = datetime.datetime(month=pair[0].month, day=pair[0].day, year=pair[0].year)
        if date not in dates_dict: dates_dict[date] = pair[1]
        else: dates_dict[date] += pair[1]
    sums_list = []
    map(lambda x: sums_list.append(['%s-%s-%s' % (x.month, x.day, x.year),
        float(dates_dict[x])]), sorted(dates_dict.keys()))
    return sums_list
