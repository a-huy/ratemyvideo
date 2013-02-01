# File to store helper functions for views
import tempfile
import os
import re
import datetime
import csv

from openpyxl.workbook import Workbook
from openpyxl.writer.excel import ExcelWriter
from openpyxl.cell import get_column_letter
from django.utils.timezone import now
from django.http import HttpResponse
from django.db.models import Count
from django.core.cache import cache
import base.cache_keys as keys
import videos.models as videos_models

# Export a list of users into a .xlsx spreadsheet
def export_users_list_to_xlsx(users):
    workbook = Workbook()
    out_file = 'users-list-' + str(now())
    ws = workbook.worksheets[0]
    ws.title = 'Users List'

    # Create column labels
    headers = ['First Name', 'Middle Name', 'Last Name', 'Email', 'Age', 'Gender',
               'Location', 'Rated', 'Balance', 'Earned', 'Joined']
    for argi in xrange(1, len(headers) + 1):
        col = get_column_letter(argi)
        cell =  ws.cell('%s1' % col)
        cell.value = headers[argi - 1]
        cell.style.font.bold = True
        if argi < 4: ws.column_dimensions[col].width = 12
        elif argi == 4: ws.column_dimensions[col].width = 25
        elif argi == 7: ws.column_dimensions[col].width = 20
        else: ws.column_dimensions[col].width = 10

    for argi in xrange(2, len(users) + 2):
        name_parts = users[argi - 2].real_name.split(' ')
        users[argi - 2].first_name = name_parts[0]
        users[argi - 2].last_name = name_parts[-1]
        users[argi - 2].middle_name = name_parts[1] if len(name_parts) > 2 else ''
        users[argi - 2].joined = users[argi - 2].created_date
        users[argi - 2].gender = users[argi - 2].gender[0].upper()
        for argt in xrange(1, len(headers) + 1):
            col = get_column_letter(argt)
            attr = re.sub(r' ', '_', headers[argt - 1].lower())
            cell = ws.cell('%s%s' % (col, argi))
            cell.value = getattr(users[argi - 2], attr)
            cell.style.font.size = 8

    fd, fn = tempfile.mkstemp()
    os.close(fd)
    workbook.save(filename = fn)
    fh = open(fn, 'rb')
    resp = fh.read()
    fh.close()
    response = HttpResponse(resp, mimetype='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment; filename="%s.xlsx"' % out_file
    return response

def export_masspay_csv(users):
    out_file = 'masspay-%s' % str(now())
    response = HttpResponse(mimetype='text/csv')
    response['Content-Disposition'] = 'attachment; filename="%s.csv"' % out_file

    writer = csv.writer(response)
    for user in users:
        if not user.pp_email: continue
        writer.writerow([user.pp_email, user.balance, 'USD', 'rmv-%s' % (user.fb_id), 'Rate My Video Payout'])
    return response

# processes a user-uploaded mass payout csv file
def process_csv(in_file):
    po_dict = {}
    curr_line = ''
    for char in in_file.read():
        if char == '\r': continue
        elif char == '\n':
            parts = curr_line.split(',')
            po_dict[parts[3][4:]] = parts[1]
            curr_line = ''
        else: curr_line += char
    in_file.close()
    return po_dict if po_dict else None

# gets the number of unique users that have rated a video today
def get_daily_user_count():
    users = cache.get(keys.RMV_USERS_DAILY_ACTIVE)
    if not users:
        now_dt = now() - datetime.timedelta(hours=8) # UTC to PST
        day = datetime.datetime(year=now_dt.year, month=now_dt.month, day=now_dt.day,
            tzinfo=now_dt.tzinfo)
        users = videos_models.Rating.active.filter(created_date__gte=day).aggregate(Count('user', distinct=True))
        cache.set(keys.RMV_USERS_DAILY_ACTIVE, users, 600)
    return users['user__count']
