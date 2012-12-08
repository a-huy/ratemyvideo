from decimal import *
import datetime
import tempfile
import os
import re

from openpyxl.workbook import Workbook
from openpyxl.writer.excel import ExcelWriter
from openpyxl.cell import get_column_letter

from django.conf import settings
from django.shortcuts import render_to_response, render, redirect
from django.http import HttpResponseRedirect, HttpResponseBadRequest, Http404, \
    HttpResponse
from django.template import RequestContext
from django.contrib.auth.decorators import login_required
from django.contrib.auth import authenticate, login, logout
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.utils.timezone import now

import videos.models as videos_models
import accounts.models as accounts_models
import base.contrib

@login_required
def home(request):
    return render_to_response('home.html', { },
        context_instance=RequestContext(request))

def rmv_login(request):
    if request.POST:
        username = request.POST['username']
        password = request.POST['password']
        user = authenticate(username=username, password=password)
        if user is not None:
            if user.is_active:
                login(request, user)
                return redirect(request.POST['next'])
    return render_to_response('login.html', { }, 
        context_instance=RequestContext(request))

def rmv_logout(request):
    logout(request)
    return redirect('/')

@login_required
def list_videos(request):
    col_filter = request.GET.get('filter', 'id')
    filter_rev = True if 'rev' in request.GET else False
    videos_all = videos_models.Video.active.all()
    ratings = videos_models.Rating.active.all()
    num_per_page = request.GET.get('num', len(videos_all))
    page = request.GET.get('page')
    paginator = Paginator(videos_all, num_per_page)
    try:
        videos = paginator.page(page)
    except PageNotAnInteger: videos = paginator.page(1)
    except EmptyPage: videos = paginator.page(paginator.num_pages)

    getcontext()
    for vid in videos:
        rating_matches = filter(lambda x: x.video_id == vid.id, ratings)
        vid.count = len(rating_matches)
        if len(rating_matches) == 0: vid.avg_rating = 0
        else: vid.avg_rating = \
            round(Decimal(sum([x.rating for x in rating_matches])) / Decimal(len(rating_matches)), 3)

    videos.object_list = sorted(videos, key=lambda x: getattr(x, col_filter), reverse=filter_rev)

    context_vars = {
        'videos': videos,
        'filter': col_filter,
        'rev': filter_rev
    }
    return render_to_response('list_videos.html', context_vars,
        context_instance=RequestContext(request))

@login_required
def list_users(request):
    col_filter = request.GET.get('filter', 'id')
    filter_rev = True if 'rev' in request.GET else False
    users_all = accounts_models.User.active.all()
    num_per_page = request.GET.get('num', len(users_all))
    page = request.GET.get('page')
    paginator = Paginator(users_all, num_per_page)
    try:
        users = paginator.page(page)
    except PageNotAnInteger: users = paginator.page(1)
    except EmptyPage: videos = paginator.page(paginator.num_pages)
    inv_requests = accounts_models.InviteRequest.objects.all().order_by('-created_date')
    for user in users.object_list:
        invites = filter(lambda x: x.fb_id == user.fb_id, inv_requests)
        user.referral = invites[0].reason if invites else 'None'
        user.tslr = base.contrib.time_since_last_rating(user.pk)
        # CST to PST (this is really horrible, but it's a hotfix for the moment)
        user.created_date -= datetime.timedelta(hours=2)
    users.object_list = sorted(users, key=lambda x: getattr(x, col_filter), reverse=filter_rev)

    if 'export' in request.GET:
        print request.get_full_path()
        return export_users_list_to_xlsx(users.object_list)

    context_vars = {
        'users': users,
        'filter': col_filter,
        'rev': filter_rev,
        'total_active': len(filter(lambda x: x.rated > 0, users_all)),
        'total_payout': len(filter(lambda x: x.balance > 10, users_all)),
        'total_users': len(users_all)
    }
    return render_to_response('list_users.html', context_vars,
        context_instance=RequestContext(request))

@login_required
def user_info(request, fb_id):
    try:
        user = accounts_models.User.active.get(fb_id=fb_id)
        user.referral = accounts_models.InviteRequest.objects.filter(fb_id=fb_id)[0].reason
    except accounts_models.User.DoesNotExist:
        return redirect(list_users)
    except IndexError: user.referral = 'None'
    videos_all = videos_models.Video.active.all()
    queue = videos_models.Queue.active.filter(user_id=user.id)
    ratings = videos_models.Rating.active.filter(user_id=user.id)
    votes = videos_models.Vote.active.filter(user_id=user.id)
    acc_age = now() - user.created_date

    for vid in queue:
        video = filter(lambda x: x.id == vid.video_id, videos_all)[0]
        vid.title = video.title
        vid.yt_id = video.yt_id
        # CST to PST (this is really horrible, but it's a hotfix for the moment)
        vid.expire_date -= datetime.timedelta(hours=2)
        vid.reward = video.reward

    for entry in ratings:
        video = filter(lambda x: x.id == entry.video_id, videos_all)[0]
        entry.title = video.title
        entry.yt_id = video.yt_id
        vote = filter(lambda x: x.video_id == entry.video_id, votes)
        if vote: entry.liked = 't' if vote[0].like else 'f'

    json_vars = {
        'verified': str(user.verified),
        'fb_id': str(user.fb_id)
    }
    context_vars = {
        'user': user,
        'queue': queue,
        'ratings': ratings,
        'votes': votes,
        'acc_age': acc_age.days,
        'json_vars': json_vars
    }
    return render_to_response('user_info.html', context_vars,
        context_instance=RequestContext(request))

@login_required
def add_video(request):
    return render_to_response('add_video.html', { },
        context_instance=RequestContext(request))

@login_required
def whitelist(request):
    return render_to_response('whitelist.html', { },
        context_instance=RequestContext(request))

@login_required
def invites(request):
    inv_requests = accounts_models.InviteRequest.active.all()
    context_vars = {
        'invite_requests': inv_requests
    }
    return render_to_response('invites.html', context_vars,
        context_instance=RequestContext(request))

def export_users_list_to_xlsx(users):
    workbook = Workbook()
    out_file = r'users-list-' + str(now())
    ws = workbook.worksheets[0]
    ws.title = 'Users List'

    # Create column labels
    headers = ['First Name', 'Middle Name', 'Last Name', 'Email', 'Age', 'Gender',
               'Location', 'Rated', 'Balance', 'Earned', 'Joined']
    for argi in xrange(1, len(headers) + 1):
        col = get_column_letter(argi)
        cell =  ws.cell('%s1' % col)
        cell.value = headers[argi - 1]
        cell.style.font.bold = True
        if argi < 4: ws.column_dimensions[col].width = 12
        elif argi == 4: ws.column_dimensions[col].width = 25
        elif argi == 7: ws.column_dimensions[col].width = 20
        else: ws.column_dimensions[col].width = 10

    for argi in xrange(2, len(users) + 2):
        name_parts = users[argi - 2].real_name.split(' ')
        users[argi - 2].first_name = name_parts[0]
        users[argi - 2].last_name = name_parts[-1]
        users[argi - 2].middle_name = name_parts[1] if len(name_parts) > 2 else ''
        users[argi - 2].joined = users[argi - 2].created_date
        users[argi - 2].gender = users[argi - 2].gender[0].upper()
        for argt in xrange(1, len(headers) + 1):
            col = get_column_letter(argt)
            attr = re.sub(r' ', '_', headers[argt - 1].lower())
            cell = ws.cell('%s%s' % (col, argi))
            cell.value = getattr(users[argi - 2], attr)
            cell.style.font.size = 8

    fd, fn = tempfile.mkstemp()
    os.close(fd)
    workbook.save(filename = fn)
    fh = open(fn, 'rb')
    resp = fh.read()
    fh.close()
    response = HttpResponse(resp, mimetype='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment; filename="%s.xlsx"' % out_file
    return response

